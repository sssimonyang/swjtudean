#!/usr/bin/env python 
# -*- coding: utf-8 -*-
"""
@author: SSSimon Yang
@contact: yangjingkang@126.com
@file: tuimian.py
@time: 2019/8/24 21:34
@desc: install packages list
conda install pyopenxl lxml
excel 目录设置方式：
增加名称 X=MID(GET.WORKBOOK(1),FIND("]",GET.WORKBOOK(1))+1,100)
在 A1 输入 =HYPERLINK("#'"&INDEX(X,ROW())&"'!A1",INDEX(X,ROW())) 下拉，保存为xlsm文件
"""
import xml.etree.ElementTree as ET

from lxml import etree
from openpyxl import Workbook

from config import user_id, user_password
from main.login import login


def main(grade="2016"):
    s = login(user_id=user_id, user_password=user_password)
    res = s.get("http://jwc.swjtu.edu.cn/vatuu/AjaxXML?selectType=CollegeInfo&selectValue=teachingCollege")
    page = ET.fromstring(res.text)
    college_names = [i.text for i in page.iter("college_name")]
    college_codes = [i.text for i in page.iter("college_code")]
    params1 = {
        "selectType": "SpecialityInfo",
        "selectValue": "collegeSpeciality",
        "key2": "1"
    }
    params2 = {
        "setAction": "specialityPlan",
        "SelectType": "query",
        "B1": "查询"
    }
    wb = Workbook()
    for college_name, college_code in zip(college_names, college_codes):
        params1["key"] = college_code
        res = s.get("http://jwc.swjtu.edu.cn/vatuu/AjaxXML", params=params1)
        if "speciality_code" not in res.text:
            continue
        page = ET.fromstring(res.text)
        speciality_codes = [i.text for i in page.iter("speciality_code")]
        speciality_names = [i.text for i in page.iter("speciality_name")]
        for speciality_name, speciality_code in zip(speciality_names, speciality_codes):
            if speciality_name in ('建筑学', '城乡规划', '风景园林'):
                params2["grade"] = str(int(grade) - 1)
            else:
                params2["grade"] = grade
            results = []
            params2["college_name"] = college_name
            params2["speciality_name"] = speciality_name
            params2["college_code"] = college_code
            params2["speciality_code"] = speciality_code
            res = s.get("http://jwc.swjtu.edu.cn/vatuu/MasterAction", params=params2)
            if not res.ok:
                continue
            page = etree.HTML(res.text)
            lines = page.xpath("//table[@id='table6']/tr")
            if len(lines) == 1:
                continue
            lines.pop()
            for line in lines:
                temp = [i for i in line.getchildren()]
                result = [i.text.strip() for i in temp[:-1]]
                if temp[-1].getchildren():
                    result.append("非免研课")
                else:
                    result.append(temp[-1].text.strip())
                results.append(result)
            sheet = wb.create_sheet(f"{college_name}-{speciality_name}".replace("[", "").replace("]", ""))
            for i, result in enumerate(results):
                for j, k in enumerate(result):
                    sheet.cell(row=i + 1, column=j + 1).value = k

    wb.save(f"2020年保研课{grade}级教务查询.xlsx")


if __name__ == '__main__':
    main()
