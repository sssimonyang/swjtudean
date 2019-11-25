#!/usr/bin/env python 
# -*- coding: utf-8 -*-
"""
@author: SSSimon Yang
@contact: yangjingkang@126.com
@file: jiaowu.py
@time: 2019/8/25 11:30
@desc:
"""
import os
import re
import xml.etree.ElementTree as ElementTree
from itertools import product
from urllib.parse import urljoin

import openpyxl
import psycopg2
from lxml import etree

import config
from main.login import login

day_of_week_to_number = {
    "星期一": 1,
    "星期二": 2,
    "星期三": 3,
    "星期四": 4,
    "星期五": 5,
    "星期六": 6,
    "星期日": 7
}


def get_all_courses():
    term = "2019-2020学年上学期"
    html = "http://jwc.swjtu.edu.cn/vatuu/CourseAction"
    params = {"setAction": "queryCourseList",
              "viewType": "",
              "orderType": "teachId",
              "orderValue": "asc",
              "selectAction": "QueryAll",
              "key1": "",
              "key2": "",
              "key3": "",
              "key4": "",
              "selectTermId": "92",
              "selectTermName": "2019-2020第1学期",
              "courseType": "all",
              "selectTableType": "ThisTerm",
              "jumpPage": "1"}
    s = login(user_id=config.ly_user_id, user_password=config.ly_user_password)

    def course_page(page_number):

        params["jumpPage"] = page_number
        single_res = s.get(html, params=params)
        single_page = etree.HTML(single_res.text)
        lists = single_page.xpath("//table[@class='c-tb']/tbody/tr")
        rows = []
        for i in lists:
            number = i.xpath("td[2]/a/font/text()")[0].strip()
            code = i.xpath("td[3]/a/text()")[0].strip()
            name = i.xpath("td[4]/a/text()")[0].strip()
            credit = i.xpath("td[6]/a/text()")[0].strip()
            course_type = i.xpath("td[7]/a/text()")[0].strip()
            teacher_name_list = i.xpath("td[9]/a/text()")
            teacher_name_list = [j.strip() for j in teacher_name_list]
            teacher_name = "+".join(teacher_name_list)
            time_location_list = i.xpath("td[11]/text()")
            time_location_list = [j.strip() for j in time_location_list]
            time_location_list = [j for j in time_location_list if j]
            assert len(time_location_list) % 2 == 0, f"the length of {time_location_list} is out of control"
            time_location = "+".join(time_location_list)
            classes_list = i.xpath("td[12]/a/text()")
            classes_list = [j.strip() for j in classes_list]
            classes = "+".join(classes_list)
            status = i.xpath("td[13]/text()")[0].strip()
            number_of_people, class_capacity = status.split("/")
            campus = i.xpath("td[14]/a/text()")[0].strip()
            if len(i.xpath("td[15]/p")) == 4:
                name_list_url = i.xpath("td[15]/p[2]/a/@href")[0]
            else:
                name_list_url = i.xpath("td[15]/p[1]/a/@href")[0]
            name_list_url = urljoin(html, name_list_url)
            rows.append(
                (number, code, name, credit, course_type, teacher_name,
                 time_location, classes, number_of_people, class_capacity, campus, name_list_url, term))
        c.executemany("INSERT INTO course values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) on conflict do nothing", rows)

        print(f"{page_number} 页 {len(rows)} and {c.statusmessage}")

    res = s.get(html, params=params)
    page = etree.HTML(res.text)
    page_total = len(page.xpath("//select[@class='btn btn-page']/option"))

    for page_num in range(page_total):
        course_page(page_num + 1)
    conn.commit()
    conn.close()


def get_all_names():
    s = login(user_id=config.ly_user_id, user_password=config.ly_user_password)

    def name_single(url, course, term):
        single_res = s.get(url)
        single_page = etree.HTML(single_res.text)
        lists = single_page.xpath("//table[@id='table2']/tr")
        if not lists:
            return
        lists.pop(0)
        rows = []
        for i in lists:
            student_id = i.xpath("td[2]/text()")[0].strip()
            student_name = i.xpath("td[3]/text()")[0].strip()
            student_gender = i.xpath("td[4]/text()")[0].strip()
            student_class = i.xpath("td[5]/text()")[0].strip()
            rows.append((course, student_id, student_name, student_class, student_gender, term))
        c.executemany("INSERT INTO student_course values (%s,%s,%s,%s,%s,%s) on conflict do nothing ", rows)
        print(f"课程{course} {len(rows)} and {c.statusmessage}")

    c.execute("SELECT name_list_url,number,term FROM course order by number ")
    results = c.fetchall()
    for num, line in enumerate(results):
        name_single(*line)
        print(num + 1, line[1])
        if num + 1 % 100 == 0:
            conn.commit()
    conn.commit()
    conn.close()


def single_student_course_time_location(student_id):
    c.execute(f"""select c.name, c.time_location,sc.student_name
                    from course c,student_course sc
                    where c.number = sc.course and sc.student_id = '{student_id}'""")
    courses = c.fetchall()

    course_time_locations = []
    for course in courses:
        if course[1] == '':
            pass
        temp = course[1].split("+")
        course_time_locations.append([course[0], temp[0], temp[1]])
        if len(temp) == 4:
            course_time_locations.append([course[0], temp[2], temp[3]])

    results = []
    for course_time_location in course_time_locations:
        course, time, location = course_time_location
        temp = time.split(" ")
        if len(temp) == 1:
            weeks = temp[0]
            day_of_week = 0  # user zero to imply unknown day_of_week
            time_of_day = 0
        else:
            assert len(temp) == 3, f"the length of {temp} is out of control"
            weeks = temp[0]
            day_of_week = day_of_week_to_number[temp[1]]
            nums = [int(i) for i in temp[2].strip("节").split("-")]
            time_of_day = list(range(nums[0], nums[1] + 1))
        results.append([course, weeks, day_of_week, time_of_day, location])
    return results


def single_student_write_to_excel(ws, results):
    number_of_unknown = 0
    for result in results:
        course, weeks, day_of_week, time_of_days, location = result
        if day_of_week == 0:
            ws.cell(row=15 + number_of_unknown, column=2).value = " ".join([course, weeks, "时间未知", location])
            number_of_unknown = number_of_unknown + 1
        else:
            for time_of_day in time_of_days:
                ws.cell(row=time_of_day + 1, column=day_of_week + 1).value = " ".join([course, weeks, location])


def get_time_to_excel(student_ids, student_names, job_name, detailed=False):
    people_of_all = len(student_ids)
    results = {}
    for student_id, student_name in zip(student_ids, student_names):
        results[student_name] = single_student_course_time_location(student_id)
    times_for_people = {}
    for student_name, result in results.items():
        single_times = []
        for i in result:
            if i[2] == 0:
                pass
            else:
                for j in i[3]:
                    single_times.append((i[2], j))
        times_for_people[student_name] = single_times

    times_for_time = {}
    for i, j in product(range(1, 6), range(1, 14)):
        times_for_time[(i, j)] = [people for people, times in times_for_people.items()
                                  if (i, j) in times]

    wb = openpyxl.load_workbook("模板.xlsx")
    ws1 = wb["空闲人数"]
    ws2 = wb["非空闲人数"]

    for i, j in product(range(1, 6), range(1, 14)):
        ws1.cell(column=i + 1, row=j + 1).value = people_of_all - len(times_for_time[(i, j)])
        ws2.cell(column=i + 1, row=j + 1).value = len(times_for_time[(i, j)])
    ws1.cell(row=15, column=1).value = f"共有 {people_of_all} 人"
    ws2.cell(row=15, column=1).value = f"共有 {people_of_all} 人"
    ws3 = wb["非空闲名单"]
    line = 1
    for time, value in times_for_time.items():
        if value:
            line = line + 1
            ws3.cell(line, 1).value = time[0]
            ws3.cell(line, 2).value = time[1]
            ws3.cell(line, 3).value = " ".join(value)

    if not detailed:
        return
    for student_name, result in results.items():
        ws = wb.copy_worksheet(wb["模板"])
        ws.title = student_name
        single_student_write_to_excel(ws, result)

    wb.active = wb.worksheets[0]
    wb.remove(wb["模板"])
    wb.save(f"{job_name}.xlsx")


def check(student_ids, student_names):
    assert len(student_ids) == len(student_names)
    student_ids = tuple(student_ids)
    c.execute(f"""select distinct sc.student_id,sc.student_name
                    from student_course sc
                    where student_id in {student_ids}""")
    ids_and_names = c.fetchall()
    if len(ids_and_names) == 0:
        return False
    for i, j in zip(student_ids, student_names):
        if (i, j) not in ids_and_names:
            print(i, j)
            return False
    return True


def process_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    student_ids = []
    student_names = []
    for i, j in ws.iter_rows(2):
        student_names.append(str(i.value))
        student_ids.append(str(j.value))
    return student_ids, student_names


def main_excel(path, job_name="", detailed=False):
    try:
        _ids, _names = process_excel(path)
    except ValueError:
        print("excel格式错误")
        return 1
    if not check(_ids, _names):
        print("student ids and student names are not correspond")
        return 2
    if not job_name:
        job_name, _ = os.path.splitext(os.path.split(path)[1])
        job_name = job_name + "空闲时间统计"
    get_time_to_excel(_ids, _names, job_name=job_name, detailed=detailed)
    return 0


def main_ids_and_names(ids_and_names, job_name='', detailed=False):
    _ids, _names = ids_and_names
    if not check(_ids, _names):
        print("student ids and student names are not correspond")
        return 2
    if not job_name:
        job_name = "空闲时间统计"
    get_time_to_excel(_ids, _names, job_name=job_name, detailed=detailed)
    return 0


def get_all_classes():
    s = login(user_id=config.ly_user_id, user_password=config.ly_user_password)
    res = s.get("http://jwc.swjtu.edu.cn/vatuu/AjaxXML?selectType=CollegeInfo&selectValue=allCollege")
    page = ElementTree.fromstring(res.text)
    college_names = [i.text for i in page.iter("college_name")]
    college_codes = [i.text for i in page.iter("college_code")]
    url = "http://jwc.swjtu.edu.cn/vatuu/PublicInfoQueryAction"
    data = {
        "setAction": "queryClass",
        "collegeCode": "16",
        "btn1": "执行查询"
    }
    for college_name, college_code in list(zip(college_names, college_codes))[2:3]:
        print(college_name)
        data["collegeCode"] = college_code
        res = s.post(url, data=data)
        page = etree.HTML(res.text)
        trs = page.xpath("//table[@class='table_gray']/tr")
        trs = trs[2:]
        if len(trs) <= 1:
            continue
        class_codes = [i.xpath("td[2]/text()")[0].strip() for i in trs]
        class_names = [i.xpath("td[3]/text()")[0].strip() for i in trs]
        _college_names = [i.xpath("td[4]/text()")[0].strip() for i in trs]
        major_names = [i.xpath("td[5]/text()")[0].strip() for i in trs]
        grade = [i.xpath("td[6]/text()")[0].strip() for i in trs]
        class_num = [i.xpath("td[7]/text()")[0].strip() for i in trs]
        c.executemany("INSERT INTO class values (%s,%s,%s,%s,%s,%s) ",
                      zip(class_codes, class_names, _college_names, major_names, grade, class_num))
        conn.commit()


def students():
    c.execute("""select distinct student_id,student_name,student_class,student_gender,major_name,college_name,grade
                    from student_course,class
                    where student_class = class_name""")
    results = [(*i[0:6], re.match(r"(20)?\d{2}", i[0]).group(), str(i[-1]),
                "true" if re.search(r".*\[.*\].*", i[2]) else "false") for i in c]
    c.executemany("INSERT INTO student values (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                  results)
    conn.commit()
    c.execute("""select distinct student_id,student_name,student_class,student_gender
                        from student_course
                        where student_class not in (select class_name from class)""")
    results = [(*i, "null", "null", re.match(r"(20)?\d{2}", i[0]).group(), "0",
                "true" if re.search(r".*\[.*\].*", i[2]) else "false") for i in c]
    # enrol_grade change manually
    c.executemany("INSERT INTO student values (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                  results)

    conn.commit()


if __name__ == '__main__':
    conn = psycopg2.connect(dbname=config.dbname, user=config.user, password=config.password, host=config.host,
                            port=config.port)
    c = conn.cursor()

    # get_all_courses()
    # get_all_classes()
    # get_all_names()
    # students()

    # main_excel(path="***.xlsx", job_name="***", detailed=True)

    # names = ['***', '***']
    # ids = ['**********',
    #        '**********']
    # main_ids_and_names((ids, names), detailed=True, job_name="temp")

    conn.close()
